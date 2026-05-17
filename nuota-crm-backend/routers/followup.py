# 客户跟进记录 + 数据导出 API
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import desc
from sqlalchemy.orm import Session
import io

from database import get_db
from models import Member, Payment, AdminUser, FollowUp, Consultant, Service, ServicePackage, PackageConsumption
from utils.auth import get_current_admin
from utils.helpers import ok, to_dict

router = APIRouter(prefix="/admin", tags=["followup-export"])

# ─────────────────────────────────────────────
# 跟进状态映射
# ─────────────────────────────────────────────
STATUS_LABEL = {
    "intention": "意向客户",
    "following": "跟进中",
    "closed": "已成交",
    "lost": "已流失",
    "silent": "沉默客户",
}

TYPE_LABEL = {
    "note": "备注",
    "call": "电话",
    "visit": "拜访",
    "wechat": "微信",
}


# ─────────────────────────────────────────────
# 跟进记录 CRUD
# ─────────────────────────────────────────────
class FollowUpIn(BaseModel):
    content: str
    status: Optional[str] = "following"
    follow_type: Optional[str] = "note"
    next_follow_date: Optional[datetime] = None


@router.get("/members/{mid}/followups")
def list_followups(
    mid: int,
    db: Session = Depends(get_db),
    _: AdminUser = Depends(get_current_admin),
):
    """获取某学员的所有跟进记录"""
    rows = (
        db.query(FollowUp)
        .filter(FollowUp.member_id == mid)
        .order_by(desc(FollowUp.created_at))
        .all()
    )
    return ok([{
        "id": r.id,
        "content": r.content,
        "status": r.status,
        "status_label": STATUS_LABEL.get(r.status, r.status),
        "follow_type": r.follow_type,
        "type_label": TYPE_LABEL.get(r.follow_type, r.follow_type),
        "next_follow_date": str(r.next_follow_date)[:16] if r.next_follow_date else None,
        "admin_name": r.admin_name,
        "created_at": str(r.created_at)[:16] if r.created_at else "",
    } for r in rows])


@router.post("/members/{mid}/followups")
def add_followup(
    mid: int,
    body: FollowUpIn,
    db: Session = Depends(get_db),
    current: AdminUser = Depends(get_current_admin),
):
    """新增跟进记录，同时更新学员跟进状态"""
    member = db.query(Member).filter(Member.id == mid).first()
    if not member:
        from fastapi import HTTPException
        raise HTTPException(404, "学员不存在")

    f = FollowUp(
        member_id=mid,
        admin_id=current.id,
        admin_name=current.real_name or current.username,
        content=body.content,
        status=body.status,
        follow_type=body.follow_type,
        next_follow_date=body.next_follow_date,
    )
    db.add(f)

    # 同步更新学员状态字段（用 role 字段临时复用，后续可加专属字段）
    # 此处不强制覆盖，由前端逻辑控制
    db.commit()
    db.refresh(f)
    return ok({
        "id": f.id,
        "content": f.content,
        "status": f.status,
        "status_label": STATUS_LABEL.get(f.status, f.status),
        "follow_type": f.follow_type,
        "type_label": TYPE_LABEL.get(f.follow_type, f.follow_type),
        "next_follow_date": str(f.next_follow_date)[:16] if f.next_follow_date else None,
        "admin_name": f.admin_name,
        "created_at": str(f.created_at)[:16] if f.created_at else "",
    })


@router.delete("/followups/{fid}")
def delete_followup(
    fid: int,
    db: Session = Depends(get_db),
    _: AdminUser = Depends(get_current_admin),
):
    """删除跟进记录"""
    f = db.query(FollowUp).filter(FollowUp.id == fid).first()
    if not f:
        from fastapi import HTTPException
        raise HTTPException(404, "记录不存在")
    db.delete(f)
    db.commit()
    return ok({"msg": "已删除"})


# ─────────────────────────────────────────────
# 数据导出（Excel）
# ─────────────────────────────────────────────
@router.get("/export/members")
def export_members(
    db: Session = Depends(get_db),
    _: AdminUser = Depends(get_current_admin),
):
    """导出学员列表 Excel"""
    try:
        import openpyxl
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "服务端缺少 openpyxl，请联系管理员安装")

    members = db.query(Member).order_by(Member.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学员列表"

    headers = ["学员编号", "姓名", "手机号", "企业名称", "城市", "角色",
               "会员类型", "会员等级", "入学日期", "到期日期", "状态",
               "年消费(元)", "推荐码", "注册时间"]
    ws.append(headers)

    type_map = {"trial": "体验会员", "annual": "年费会员", "vip": "VIP"}
    tier_map = {"primary": "小学生", "junior": "初中生", "senior": "高中生",
                "college": "大学生", "teacher": "老师"}
    status_map = {"active": "正常", "expired": "已到期", "frozen": "已冻结"}

    for m in members:
        ws.append([
            m.member_no or "",
            m.name or "",
            m.phone or "",
            m.enterprise_name or "",
            m.city or "",
            m.role or "",
            type_map.get(m.member_type, m.member_type or ""),
            tier_map.get(m.member_tier, m.member_tier or ""),
            str(m.enroll_date) if m.enroll_date else "",
            str(m.expire_date) if m.expire_date else "",
            status_map.get(m.status, m.status or ""),
            float(m.annual_spending) if m.annual_spending else 0,
            m.referral_code or "",
            str(m.created_at)[:16] if m.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_members = f"members_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_members}"'},
    )


@router.get("/export/payments")
def export_payments(
    db: Session = Depends(get_db),
    _: AdminUser = Depends(get_current_admin),
):
    """导出收款明细 Excel"""
    try:
        import openpyxl
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "服务端缺少 openpyxl，请联系管理员安装")

    rows = (
        db.query(Payment, Member)
        .outerjoin(Member, Payment.member_id == Member.id)
        .order_by(desc(Payment.id))
        .limit(5000)
        .all()
    )

    # 预加载关联数据
    consultant_ids = list({p.consultant_id for p, m in rows if p.consultant_id})
    consultants_map = {c.id: c for c in db.query(Consultant).filter(Consultant.id.in_(consultant_ids)).all()} if consultant_ids else {}

    pkg_ids = list({p.package_id for p, m in rows if getattr(p, 'package_id', None)})
    packages_map = {pk.id: pk for pk in db.query(ServicePackage).filter(ServicePackage.id.in_(pkg_ids)).all()} if pkg_ids else {}

    svc_ids = list({p.service_id for p, m in rows if getattr(p, 'service_id', None)})
    services_map = {s.id: s for s in db.query(Service).filter(Service.id.in_(svc_ids)).all()} if svc_ids else {}

    # 消耗明细
    all_consumptions = db.query(PackageConsumption).filter(
        PackageConsumption.package_id.in_(pkg_ids)
    ).order_by(PackageConsumption.visit_number).all() if pkg_ids else []
    consumption_by_pkg = {}
    for cc in all_consumptions:
        consumption_by_pkg.setdefault(cc.package_id, []).append(cc)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "收款明细"

    headers = ["收款ID", "学员编号", "学员姓名", "企业名称", "手机号",
               "归属老师", "合作项目", "金额(元)", "欠款(元)",
               "付款类型", "付款方式", "付款状态", "付款时间",
               "服务次数", "已用次数", "剩余次数", "每次扣费",
               "备注", "录入时间"]
    ws.append(headers)

    pay_type_map = {"trial": "体验费", "annual": "年费", "single": "单次"}
    pay_status_map = {"pending": "待付款", "paid": "已结清", "partial": "分期中", "refunded": "已退款"}
    pay_method_map = {"company_account": "对公账户", "private_account": "私户转账",
                      "wecom": "企业微信", "wechat_proxy": "微信代收"}

    for p, m in rows:
        consultant = consultants_map.get(p.consultant_id) if p.consultant_id else None
        pkg = packages_map.get(p.package_id) if getattr(p, 'package_id', None) else None
        svc = services_map.get(p.service_id) if getattr(p, 'service_id', None) else None

        ws.append([
            p.id,
            m.member_no if m else "",
            m.name if m else "",
            getattr(m, 'enterprise_name', '') if m else "",
            m.phone if m else "",
            consultant.name if consultant else "",
            svc.name if svc else "",
            float(p.amount) if p.amount else 0,
            float(p.debt_amount) if p.debt_amount else 0,
            pay_type_map.get(p.pay_type, p.pay_type or ""),
            pay_method_map.get(p.pay_method, p.pay_method or ""),
            pay_status_map.get(p.pay_status, p.pay_status or ""),
            str(p.pay_time)[:16] if p.pay_time else "",
            pkg.total_times if pkg else "",
            pkg.used_times if pkg else "",
            (pkg.total_times - pkg.used_times) if pkg else "",
            float(pkg.per_time_fee) if pkg and pkg.per_time_fee else "",
            p.remark or "",
            str(p.created_at)[:16] if p.created_at else "",
        ])

    # Sheet 2: 消耗明细
    if all_consumptions:
        ws2 = wb.create_sheet("消耗明细")
        ws2.append(["套餐ID", "期数", "服务内容", "分类", "扣费(元)",
                    "下店日期", "天数", "门店", "执案老师", "助理",
                    "满意度", "评价", "执案摘要"])
        for cc in all_consumptions:
            ws2.append([
                cc.package_id,
                cc.visit_number,
                cc.service_name or "",
                cc.service_category or "",
                float(cc.deducted_amount) if cc.deducted_amount else 0,
                str(cc.appoint_date) if cc.appoint_date else "",
                cc.duration_days or "",
                cc.store_name or "",
                cc.consultant_name or "",
                cc.assistant_name or "",
                cc.rating or "",
                cc.rating_comment or "",
                cc.summary or "",
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_payments = f"payments_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_payments}"'},
    )
